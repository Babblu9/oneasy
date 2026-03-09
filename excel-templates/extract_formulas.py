"""
Extract ALL cell-to-cell formulas from the Docty Healthcare Business Plan Excel file.
Outputs:
  1. Raw formula audit (every cell with a formula)
  2. Cross-sheet dependency map
  3. Grouped by sheet with cell references
"""
from openpyxl import load_workbook
import re
import json

FILE = "Docty Healthcare - Business Plan.xlsx"

wb = load_workbook(FILE, data_only=False)

print("=" * 80)
print("SHEETS IN WORKBOOK:")
print("=" * 80)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s} — {ws.max_row} rows × {ws.max_column} cols")

print("\n" + "=" * 80)
print("STEP 1: RAW FORMULA AUDIT — Every cell containing a formula")
print("=" * 80)

all_formulas = []  # (sheet, cell, formula)
cross_sheet_refs = []  # formulas that reference other sheets

for sheet in wb.worksheets:
    sheet_formulas = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value
                entry = {
                    "sheet": sheet.title,
                    "cell": cell.coordinate,
                    "formula": formula
                }
                all_formulas.append(entry)
                sheet_formulas.append(entry)
                
                # Check for cross-sheet references (SheetName!CellRef pattern)
                cross_refs = re.findall(r"'?([^'!]+)'?![A-Z$]+[0-9$]+", formula)
                if cross_refs:
                    for ref_sheet in cross_refs:
                        cross_sheet_refs.append({
                            "from_sheet": sheet.title,
                            "from_cell": cell.coordinate,
                            "to_sheet": ref_sheet,
                            "formula": formula
                        })
    
    if sheet_formulas:
        print(f"\n--- Sheet: {sheet.title} ({len(sheet_formulas)} formulas) ---")
        for entry in sheet_formulas:
            print(f"  {entry['cell']} = {entry['formula']}")

print(f"\n\nTOTAL FORMULAS: {len(all_formulas)}")

print("\n" + "=" * 80)
print("STEP 2: CROSS-SHEET DEPENDENCIES")
print("=" * 80)

if cross_sheet_refs:
    for ref in cross_sheet_refs:
        print(f"  {ref['from_sheet']}!{ref['from_cell']} → references {ref['to_sheet']}")
        print(f"    Formula: {ref['formula']}")
else:
    print("  No cross-sheet references found.")

# Also dump named ranges / defined names
print("\n" + "=" * 80)
print("DEFINED NAMES / NAMED RANGES")
print("=" * 80)
if wb.defined_names:
    for name in wb.defined_names.definedName:
        print(f"  {name.name} = {name.attr_text}")
else:
    print("  None found.")

# Save raw output to JSON for further processing
output = {
    "sheets": wb.sheetnames,
    "total_formulas": len(all_formulas),
    "formulas": all_formulas,
    "cross_sheet_refs": cross_sheet_refs
}

with open("info/raw_formulas.json", "w") as f:
    json.dump(output, f, indent=2)

print(f"\n\nRaw formulas saved to info/raw_formulas.json")

# Also extract header rows for context mapping
print("\n" + "=" * 80)
print("HEADER/LABEL CONTEXT (first 10 rows of each sheet)")
print("=" * 80)

for sheet in wb.worksheets:
    print(f"\n--- Sheet: {sheet.title} ---")
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=False), 1):
        cells = []
        for cell in row:
            if cell.value is not None:
                val = str(cell.value)[:50]
                cells.append(f"{cell.coordinate}={val}")
        if cells:
            print(f"  Row {row_idx}: {' | '.join(cells)}")
